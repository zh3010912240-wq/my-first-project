import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ==================== 配置项 ====================
HTML_FILE_NAME = "DetailedStatement.htm"  # 您的htm文件名
OUTPUT_EXCEL_NAME = "交易复盘主表.xlsx"     # 生成的Excel文件名
# ================================================

def parse_html_statement(file_path):
    """解析MT4/MT5 HTML报告中的所有交易记录"""
    trades = []
    if not os.path.exists(file_path):
        return trades
        
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        html_content = f.read()
    
    tr_pattern = re.compile(r'<tr[^>]*>(.*?)</tr>', re.DOTALL)
    td_pattern = re.compile(r'<td[^>]*>(.*?)</td>', re.DOTALL)
    all_rows = tr_pattern.findall(html_content)
    
    for row_content in all_rows:
        tds = td_pattern.findall(row_content)
        td_texts = [re.sub(r'<[^<]+?>', '', td).strip() for td in tds]
        
        if len(td_texts) >= 11:
            action = td_texts[2].lower()
            if action in ['buy', 'sell']:
                try:
                    trade_time = td_texts[1]
                    open_price = float(td_texts[5])
                    sl = float(td_texts[6])
                    tp = float(td_texts[7])
                    
                    trades.append({
                        'time': trade_time,
                        'open_price': open_price,
                        'sl': sl,
                        'tp': tp
                    })
                except ValueError:
                    continue
                    
    trades.sort(key=lambda x: x['time'])
    return trades

def get_sorted_images():
    """获取当前目录下所有的图片文件，并按自然数字顺序排序"""
    valid_extensions = ('.png', '.jpg', '.jpeg', '.bmp', '.gif')
    files = [f for f in os.listdir('.') if f.lower().endswith(valid_extensions)]
    files.sort(key=lambda var: [int(x) if x.isdigit() else x for x in re.split(r'(\d+)', var)])
    return files

def main():
    if not os.path.exists(HTML_FILE_NAME):
        print(f"错误: 找不到网页文件 '{HTML_FILE_NAME}'")
        return

    print("正在解析 HTML 交易数据...")
    all_trades = parse_html_statement(HTML_FILE_NAME)
    
    # 1. 初始化或加载已有的 Excel
    headers = ["时间", "开仓价", "止损", "止盈", "图片", "入场逻辑", "情绪", "复盘"]
    existing_times = set()
    
    if os.path.exists(OUTPUT_EXCEL_NAME):
        print(f"发现已有表格 '{OUTPUT_EXCEL_NAME}'，正在准备追加新数据...")
        wb = load_workbook(OUTPUT_EXCEL_NAME)
        ws = wb.active
        # 读取第1列（时间列）所有已经存在的交易时间，防止重复添加
        for row in range(2, ws.max_row + 1):
            time_val = ws.cell(row=row, column=1).value
            if time_val:
                existing_times.add(str(time_val).strip())
    else:
        print(f"未发现历史表格，正在创建全新表格 '{OUTPUT_EXCEL_NAME}'...")
        wb = Workbook()
        ws = wb.active
        ws.title = "每日复盘"
        ws.views.sheetView[0].showGridLines = True
        ws.append(headers)
        
        # 初始表头样式
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_num in range(1, 9):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

    # 2. 过滤出真正“新做”的单子
    new_trades = [t for t in all_trades if t['time'] not in existing_times]
    
    if not new_trades:
        print("💡 没有检测到新的交易单子，表格无需更新。")
        return
        
    print(f"检测到有 {len(new_trades)} 笔新交易待追加！")
    
    # 3. 扫描图片
    images = get_sorted_images()
    
    # 样式配置
    body_font = Font(name="微软雅黑", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    # 4. 追加新数据
    start_row = ws.max_row + 1
    for idx, trade in enumerate(new_trades):
        current_row = start_row + idx
        
        # 写入新数据
        ws.cell(row=current_row, column=1, value=trade['time'])
        ws.cell(row=current_row, column=2, value=trade['open_price'])
        ws.cell(row=current_row, column=3, value=trade['sl'])
        ws.cell(row=current_row, column=4, value=trade['tp'])
        
        # 初始化样式
        for col_num in range(1, 9):
            c = ws.cell(row=current_row, column=col_num)
            c.font = body_font
            c.border = thin_border
            if col_num in [1, 2, 3, 4]:
                c.alignment = center_align
                if col_num in [2, 3, 4]:
                    c.number_format = '#,##0.00'
            else:
                c.alignment = left_align
                
        # 设置行高以容纳图片
        ws.row_dimensions[current_row].height = 110
        
        # 自动配对并插入新图片
        # 比如：当前目录下有5张图，这次新加了2个单子，它会把最后2张图塞进这2个新单子里
        img_index = len(images) - len(new_trades) + idx
        if 0 <= img_index < len(images):
            img_path = images[img_index]
            try:
                img = OpenpyxlImage(img_path)
                max_width, max_height = 180, 140
                ratio = min(max_width / img.width, max_height / img.height)
                img.width = int(img.width * ratio)
                img.height = int(img.height * ratio)
                
                ws.add_image(img, f"E{current_row}")
            except Exception as e:
                ws.cell(row=current_row, column=5, value=f"图片加载失败")

    # 统一调整列宽
    column_widths = {'A': 22, 'B': 12, 'C': 12, 'D': 12, 'E': 26, 'F': 30, 'G': 15, 'H': 40}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 保存
    wb.save(OUTPUT_EXCEL_NAME)
    print(f"成功！已安全追加 {len(new_trades)} 笔新交易到 '{OUTPUT_EXCEL_NAME}' 的末尾。")

if __name__ == "__main__":
    main()